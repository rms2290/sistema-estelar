"""
Utilitários para geração de relatórios em PDF e Excel (Totalizador por Estado).
"""
import io
from datetime import datetime
from decimal import Decimal

from django.http import HttpResponse


def format_brazilian_currency(value):
    """Formata valor como moeda brasileira."""
    if value is None:
        return 'R$ 0,00'
    try:
        float_value = float(value)
        formatted = f"{float_value:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        return f"R$ {formatted}"
    except (ValueError, TypeError):
        return str(value)


def gerar_relatorio_pdf_totalizador_estado(resultados, data_inicial, data_final, total_geral, total_seguro_geral):
    """
    Gera relatório PDF para Totalizador por Estado.
    data_inicial e data_final devem ser objetos date.
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#1a237e'),
    )
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=12,
        spaceAfter=20,
        alignment=TA_CENTER,
        textColor=colors.grey,
    )
    para_style = ParagraphStyle(
        'CustomPara',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=12,
    )

    story = []
    story.append(Paragraph("RELATÓRIO - TOTALIZADOR POR ESTADO", title_style))
    story.append(Spacer(1, 12))

    data_ini = data_inicial.strftime('%d/%m/%Y') if hasattr(data_inicial, 'strftime') else str(data_inicial)
    data_fim = data_final.strftime('%d/%m/%Y') if hasattr(data_final, 'strftime') else str(data_final)
    story.append(Paragraph(f"Período: {data_ini} a {data_fim}", subtitle_style))
    story.append(Spacer(1, 20))
    story.append(Paragraph(f"Relatório gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}", para_style))
    story.append(Spacer(1, 20))

    story.append(Paragraph("RESUMO EXECUTIVO", styles['Heading2']))
    story.append(Spacer(1, 10))
    resumo_data = [
        ['Total de Estados', str(len(resultados))],
        ['Valor Total', format_brazilian_currency(total_geral)],
        ['Seguro Total', format_brazilian_currency(total_seguro_geral)],
    ]
    resumo_table = Table(resumo_data, colWidths=[8 * cm, 6 * cm])
    resumo_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    story.append(resumo_table)
    story.append(Spacer(1, 30))

    story.append(Paragraph("DETALHAMENTO POR ESTADO", styles['Heading2']))
    story.append(Spacer(1, 10))
    table_data = [['ESTADO', 'QTD. ROMANEIOS', 'VALOR TOTAL', '% SEGURO', 'VALOR SEGURO']]
    for r in resultados:
        pct = f"{float(r['percentual_seguro']):.2f}%" if r.get('percentual_seguro') is not None else '0,00%'
        table_data.append([
            f"{r['nome_estado']} ({r['estado']})",
            str(r['quantidade_romaneios']),
            format_brazilian_currency(r['total_valor']),
            pct,
            format_brazilian_currency(r['valor_seguro']),
        ])
    table = Table(table_data, colWidths=[4 * cm, 3 * cm, 3 * cm, 2.5 * cm, 3 * cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    story.append(table)
    story.append(Spacer(1, 20))
    story.append(Paragraph("Sistema Estelar - Relatório de Totalizador por Estado", para_style))

    doc.build(story)
    pdf_content = buffer.getvalue()
    buffer.close()
    return pdf_content


def gerar_relatorio_excel_totalizador_estado(resultados, data_inicial, data_final, total_geral, total_seguro_geral):
    """
    Gera relatório Excel para Totalizador por Estado.
    data_inicial e data_final devem ser objetos date.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Totalizador por Estado"

    title_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    data_font = Font(name='Arial', size=10)
    title_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    data_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    right_alignment = Alignment(horizontal='right', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    data_ini = data_inicial.strftime('%d/%m/%Y') if hasattr(data_inicial, 'strftime') else str(data_inicial)
    data_fim = data_final.strftime('%d/%m/%Y') if hasattr(data_final, 'strftime') else str(data_final)

    ws.merge_cells('A1:F1')
    ws['A1'] = 'RELATÓRIO - TOTALIZADOR POR ESTADO'
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = center_alignment

    ws.merge_cells('A2:F2')
    ws['A2'] = f'Período: {data_ini} a {data_fim}'
    ws['A2'].font = data_font
    ws['A2'].alignment = center_alignment

    ws.merge_cells('A3:F3')
    ws['A3'] = f'Relatório gerado em: {datetime.now().strftime("%d/%m/%Y às %H:%M")}'
    ws['A3'].font = data_font
    ws['A3'].alignment = center_alignment

    ws['A5'] = 'RESUMO EXECUTIVO'
    ws['A5'].font = header_font
    ws['A5'].fill = header_fill
    ws['A5'].alignment = left_alignment
    ws['A6'] = 'Total de Estados:'
    ws['B6'] = len(resultados)
    ws['A7'] = 'Valor Total:'
    ws['B7'] = format_brazilian_currency(total_geral)
    ws['A8'] = 'Seguro Total:'
    ws['B8'] = format_brazilian_currency(total_seguro_geral)
    for row in range(6, 9):
        ws[f'A{row}'].font = data_font
        ws[f'B{row}'].font = data_font
        ws[f'B{row}'].alignment = right_alignment

    headers = ['ESTADO', 'QTD. ROMANEIOS', 'VALOR TOTAL', '% SEGURO', 'VALOR SEGURO']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    for row, resultado in enumerate(resultados, 11):
        pct = f"{float(resultado.get('percentual_seguro', 0)):.2f}%" if resultado.get('percentual_seguro') is not None else '0,00%'
        ws.cell(row=row, column=1, value=f"{resultado['nome_estado']} ({resultado['estado']})")
        ws.cell(row=row, column=2, value=resultado['quantidade_romaneios'])
        ws.cell(row=row, column=3, value=format_brazilian_currency(resultado['total_valor']))
        ws.cell(row=row, column=4, value=pct)
        ws.cell(row=row, column=5, value=format_brazilian_currency(resultado['valor_seguro']))
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.border = thin_border
            if row % 2 == 0:
                cell.fill = data_fill
            cell.alignment = left_alignment if col == 1 else center_alignment

    for col, width in enumerate([25, 15, 18, 12, 18], 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def gerar_resposta_pdf(conteudo_pdf, nome_arquivo, inline=False):
    """Cria resposta HTTP para PDF."""
    response = HttpResponse(content_type='application/pdf')
    disposition = 'inline' if inline else 'attachment'
    response['Content-Disposition'] = f'{disposition}; filename="{nome_arquivo}"'
    response.write(conteudo_pdf)
    return response


def gerar_resposta_excel(conteudo_excel, nome_arquivo):
    """Cria resposta HTTP para Excel."""
    response = HttpResponse(
        conteudo_excel,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
    return response


def gerar_relatorio_pdf_cobranca_carregamento(cobranca):
    """Gera PDF detalhado de uma cobrança de carregamento."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(f"COBRANÇA DE CARREGAMENTO #{cobranca.id}", styles['Title']))
    story.append(Spacer(1, 12))
    story.append(Paragraph(f"Cliente: {cobranca.cliente.razao_social}", styles['Normal']))
    story.append(Paragraph(f"Status: {cobranca.status}", styles['Normal']))
    story.append(Paragraph(f"Criado em: {cobranca.criado_em.strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    story.append(Spacer(1, 12))

    valores = [
        ["Campo", "Valor"],
        ["Valor carregamento", format_brazilian_currency(cobranca.valor_carregamento)],
        ["Valor CTE/Manifesto", format_brazilian_currency(cobranca.valor_cte_manifesto)],
        ["Valor CTE/Terceiro", format_brazilian_currency(getattr(cobranca, 'valor_cte_terceiro', 0) or 0)],
        ["Margem Estelar", format_brazilian_currency(cobranca.margem_carregamento)],
        ["Lucro CTE", format_brazilian_currency(getattr(cobranca, 'lucro_cte', 0) or 0)],
        ["Valor total", format_brazilian_currency(cobranca.valor_total)],
    ]
    tabela = Table(valores, colWidths=[7 * cm, 7 * cm])
    tabela.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    story.append(tabela)
    story.append(Spacer(1, 12))

    romaneios = list(cobranca.romaneios.all().order_by('codigo'))
    if romaneios:
        story.append(Paragraph("Romaneios vinculados", styles['Heading2']))
        rows = [["Código", "Data Emissão"]]
        for r in romaneios:
            data_emissao = r.data_emissao.strftime('%d/%m/%Y') if r.data_emissao else '-'
            rows.append([r.codigo, data_emissao])
        tabela_romaneios = Table(rows, colWidths=[7 * cm, 7 * cm])
        tabela_romaneios.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.beige]),
        ]))
        story.append(tabela_romaneios)

    doc.build(story)
    pdf_content = buffer.getvalue()
    buffer.close()
    return pdf_content


def gerar_relatorio_pdf_consolidado_cobranca(cobrancas, cliente_selecionado=None):
    """Gera PDF consolidado de cobranças de carregamento (layout paisagem)."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib.enums import TA_CENTER
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1.2 * cm,
        leftMargin=1.2 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'ConsolidadoTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=12,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#1a237e'),
    )
    subtitle_style = ParagraphStyle(
        'ConsolidadoSubtitle',
        parent=styles['Heading2'],
        fontSize=11,
        spaceAfter=8,
        alignment=TA_CENTER,
        textColor=colors.grey,
    )
    para_style = ParagraphStyle(
        'ConsolidadoPara',
        parent=styles['Normal'],
        fontSize=9,
        spaceAfter=8,
    )
    story = []

    story.append(Paragraph("RELATÓRIO CONSOLIDADO DE COBRANÇAS DE CARREGAMENTO", title_style))
    if cliente_selecionado:
        story.append(Paragraph(f"Cliente: {cliente_selecionado.razao_social}", subtitle_style))
    else:
        story.append(Paragraph("Cliente: TODOS", subtitle_style))
    story.append(Paragraph(f"Relatório gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}", para_style))
    story.append(Spacer(1, 12))

    # Resumo executivo no mesmo padrão visual dos totalizadores.
    qtd = len(cobrancas)
    rows = [[
        "ID", "Cliente", "Status", "Carregamento", "CTE/Manif.", "CTE/Terceiro", "Lucro CTE", "Total"
    ]]

    total_carregamento = Decimal('0.00')
    total_cte_manifesto = Decimal('0.00')
    total_cte_terceiro = Decimal('0.00')
    total_lucro_cte = Decimal('0.00')
    total_geral = Decimal('0.00')

    for c in cobrancas:
        valor_carregamento = c.valor_carregamento or Decimal('0.00')
        valor_cte_manifesto = c.valor_cte_manifesto or Decimal('0.00')
        valor_cte_terceiro = getattr(c, 'valor_cte_terceiro', Decimal('0.00')) or Decimal('0.00')
        lucro_cte = getattr(c, 'lucro_cte', valor_cte_manifesto - valor_cte_terceiro)
        valor_total = c.valor_total or Decimal('0.00')

        rows.append([
            str(c.id),
            c.cliente.razao_social,
            c.status,
            format_brazilian_currency(valor_carregamento),
            format_brazilian_currency(valor_cte_manifesto),
            format_brazilian_currency(valor_cte_terceiro),
            format_brazilian_currency(lucro_cte),
            format_brazilian_currency(valor_total),
        ])

        total_carregamento += valor_carregamento
        total_cte_manifesto += valor_cte_manifesto
        total_cte_terceiro += valor_cte_terceiro
        total_lucro_cte += lucro_cte
        total_geral += valor_total

    resumo_data = [
        ['Total de Cobranças', str(qtd)],
        ['Carregamento Total', format_brazilian_currency(total_carregamento)],
        ['CTE/Manifesto Total', format_brazilian_currency(total_cte_manifesto)],
        ['CTE/Terceiro Total', format_brazilian_currency(total_cte_terceiro)],
        ['Lucro CTE Total', format_brazilian_currency(total_lucro_cte)],
        ['Valor Total Geral', format_brazilian_currency(total_geral)],
    ]
    resumo_table = Table(resumo_data, colWidths=[7 * cm, 6 * cm])
    resumo_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
    ]))
    story.append(Paragraph("RESUMO EXECUTIVO", styles['Heading2']))
    story.append(Spacer(1, 8))
    story.append(resumo_table)
    story.append(Spacer(1, 12))

    rows.append([
        "TOTAL", "-", "-", format_brazilian_currency(total_carregamento),
        format_brazilian_currency(total_cte_manifesto),
        format_brazilian_currency(total_cte_terceiro),
        format_brazilian_currency(total_lucro_cte),
        format_brazilian_currency(total_geral),
    ])

    tabela = Table(
        rows,
        colWidths=[1.2 * cm, 8.2 * cm, 2.0 * cm, 3.0 * cm, 3.0 * cm, 3.0 * cm, 3.0 * cm, 3.0 * cm]
    )
    total_row = len(rows) - 1
    tabela.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a237e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('ALIGN', (0, 1), (2, -1), 'CENTER'),
        ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
        ('ROWBACKGROUNDS', (0, 1), (-1, total_row - 1), [colors.white, colors.beige]),
        ('BACKGROUND', (0, total_row), (-1, total_row), colors.HexColor('#D9E1F2')),
        ('FONTNAME', (0, total_row), (-1, total_row), 'Helvetica-Bold'),
    ]))

    story.append(Paragraph("DETALHAMENTO DAS COBRANÇAS", styles['Heading2']))
    story.append(Spacer(1, 8))
    story.append(tabela)
    doc.build(story)

    pdf_content = buffer.getvalue()
    buffer.close()
    return pdf_content
